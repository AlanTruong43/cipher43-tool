#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script tạo sample Excel file cho Cipher 43 Tool
Chạy: python3 create_sample_excel.py
"""

def create_sample_excel():
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        print("⚠ openpyxl not installed. Installing...")
        import subprocess
        subprocess.run(["pip3", "install", "openpyxl"], check=True)
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    from pathlib import Path

    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profiles"

    # Headers
    headers = ["profile_name", "username", "password", "totp_seed", "proxy", "notes"]
    ws.append(headers)

    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Sample data (5 profiles)
    sample_data = [
        ["Account_01", "user1@gmail.com", "SecurePass123!", "JBSWY3DPEHPK3PXP", "", "Twitter engagement"],
        ["Account_02", "user2@gmail.com", "SecurePass456!", "JBSWY3DPEHPK3PXP", "http://proxy1:8080", "OKX wallet import"],
        ["Account_03", "user3@gmail.com", "SecurePass789!", "JBSWY3DPEHPK3PXP", "", "Discord community"],
        ["Account_04", "user4@gmail.com", "SecurePass000!", "JBSWY3DPEHPK3PXP", "http://proxy2:8080", ""],
        ["Account_05", "user5@gmail.com", "SecurePass111!", "JBSWY3DPEHPK3PXP", "", "Test account"],
    ]

    for row in sample_data:
        ws.append(row)

    # Set column widths
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 30

    # Save file
    output_path = Path(__file__).parent / "samples" / "sample-profiles.xlsx"
    wb.save(output_path)
    print(f"✅ Created: {output_path}")
    print(f"📊 Profiles: {len(sample_data)}")
    print(f"📋 Columns: {', '.join(headers)}")


if __name__ == "__main__":
    create_sample_excel()
